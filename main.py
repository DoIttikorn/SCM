"""
    pip3 install requests  datetime Openpyxl pylint
    
"""
import json
import time
import os
import requests
import platform
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, worksheet


def Getjson(str: str) -> dict:
    r = requests.get(str)
    return r


def CreateTimeDateValue(timeValues: list) -> list:
    for index, xran in enumerate(timeValues):
        timeStamp = xran[0] // 1000
        dt_object = datetime.fromtimestamp(timeStamp)
        timeValues[index][0] = dt_object
    return timeValues


def IsOS():
    nameOS = platform.system()
    if nameOS == 'Wondows':
        return True
    elif nameOS == 'Linux':
        return False


def CheckOSAndFindAddressDesktop() -> str:
    name = platform.system()
    if name == 'Windows':
        DesktopAddress = os.path.join(os.path.join(
            os.environ['USERPROFILE']), 'Desktop')
    if name == 'Linux':
        DesktopAddress = os.path.join(
            os.path.join(os.environ['HOME']), 'Desktop')
    return DesktopAddress


def SaveValuesExcel(result: dict, address: str, filename: str):
    wb = Workbook()
    sheet = wb.active
    for index, item in enumerate(result):
        ws = wb.create_sheet(item)
        dataInExcel = result[item]
        
        cellA1 = ws.cell(row=1, column=1)
        cellA1.value = "Date and Time"
        cellA1.font = cellA1.font.copy(bold=True)
        cellA2 = ws.cell(row=1, column=2)
        if index == 0:
            cellA2.value = "Lux"
        elif index == 1:
            cellA2.value = "Celsius"
        cellA2.font = cellA2.font.copy(bold=True)
        for i in dataInExcel:

            ws.append(i)
        # print("index  : ",index,"su : ",item,sep="  ")

        # print(result[item])
    p = Path(address, filename)
    wb.save(p)


if __name__ == "__main__":
    t0 = time.time()

    fileName = 'result.xlsx'
    str = 'https://api.netpie.io/feed/Rapoo?apikey=8YzoLCrqTRZXLv4YVW5IvmGJoEgXByuI&granularity=1minutes&since=1year&fbclid=IwAR2JA_UpUFyW_dvOpxG1s7QdB021EzpxXa7EDlM5vgnSwmBaN1ueOuo0DjQ'
    # Get text in String
    StringInformationWeb = Getjson(str)
    # เป็นค่าภายในให้เป็น ่json ใน python อยู่ในรูป dict
    jsonContent = StringInformationWeb.json()
    # เลือกส่วนที่เป็นข้อมูล
    data = jsonContent.get('data')
    # หาที่อยู่ desktop เพื่อที่จะบันทึก excel
    desktopAddress = CheckOSAndFindAddressDesktop()
    itemDict = {}
    # วนลูปเพื่อเอาค่าแต่ละ ข้อมูลออกมา เช่น อุณหภูมิ ,ความดัน
    for target, typeData in enumerate(data):
        # ค่าตัวเลขจะอยู่ใน values
        values = typeData.get('values')
        #  ชื่อคอลัมเอาไว้ใส่ใน excel
        columnName = typeData.get('attr')
        #  เอาค่าเวลากับค่าผลลัพธ์ส่งไปเป็น list เป็น timestamp เป็นวันเวลาปกติ เก้บในตัวแปร succ
        succ = CreateTimeDateValue(values)
        itemDict[columnName] = succ
    # ทำการ save ค่าลงใน excel
    SaveValuesExcel(itemDict, desktopAddress, fileName)
    t1 = time.time()
    total = t1-t0
    print("\n\n Performance python file : ", total)
